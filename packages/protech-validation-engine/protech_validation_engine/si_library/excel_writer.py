"""Excel workbook generation for SI Library Audit results."""

from __future__ import annotations

from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from .audit_engine import (
    MATCH_EXACT,
    MATCH_MISSING,
    MATCH_POTENTIAL_MISMATCH,
    MATCH_SPLIT_NAMING,
    MATCH_SPLIT_PRESENT,
    AuditResult,
)
from .config import AuditSettings

try:
    from .executive_intelligence import ExecutivePackage, build_executive_package
except ImportError:
    ExecutivePackage = None  # type: ignore
    build_executive_package = None  # type: ignore

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F5F99", end_color="1F5F99", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14, color="0F2F57")
SECTION_FONT = Font(bold=True, size=12, color="0F2F57")


def _autosize_columns(ws, max_width: int = 60) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_length + 2, 12), max_width)


def _write_dataframe(ws, df: pd.DataFrame, start_row: int = 1, header: bool = True) -> int:
    for row_idx, row in enumerate(
        dataframe_to_rows(df, index=False, header=header),
        start=start_row,
    ):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    if header and not df.empty:
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return start_row + len(df) + (1 if header else 0)


def _apply_audit_formatting(ws, status_col: int, owner_col: int, last_row: int) -> None:
    if last_row < 2:
        return

    status_letter = get_column_letter(status_col)
    owner_letter = get_column_letter(owner_col)
    status_range = f"{status_letter}2:{status_letter}{last_row}"
    owner_range = f"{owner_letter}2:{owner_letter}{last_row}"

    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=[f'"{MATCH_EXACT}"'], fill=GREEN_FILL),
    )
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=[f'"{MATCH_SPLIT_PRESENT}"'], fill=GREEN_FILL),
    )
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=[f'"{MATCH_SPLIT_NAMING}"'], fill=YELLOW_FILL),
    )
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=[f'"{MATCH_POTENTIAL_MISMATCH}"'], fill=ORANGE_FILL),
    )
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=[f'"{MATCH_MISSING}"'], fill=RED_FILL),
    )

    ws.conditional_formatting.add(
        owner_range,
        CellIsRule(operator="equal", formula=['"Compliant"'], fill=GREEN_FILL),
    )
    ws.conditional_formatting.add(
        owner_range,
        CellIsRule(operator="equal", formula=['"Compliant / PCS Review"'], fill=YELLOW_FILL),
    )
    ws.conditional_formatting.add(
        owner_range,
        CellIsRule(operator="equal", formula=['"SI Naming Review"'], fill=YELLOW_FILL),
    )
    ws.conditional_formatting.add(
        owner_range,
        CellIsRule(operator="equal", formula=['"Needs Human Review"'], fill=ORANGE_FILL),
    )
    ws.conditional_formatting.add(
        owner_range,
        CellIsRule(operator="equal", formula=['"SI / Export Gap Review"'], fill=RED_FILL),
    )


def _compliance_tone(
    value: float,
    settings: AuditSettings | None = None,
) -> PatternFill:
    excellent = settings.compliance_threshold_excellent if settings else 90.0
    acceptable = settings.compliance_threshold_acceptable if settings else 70.0
    if value >= excellent:
        return GREEN_FILL
    if value >= acceptable:
        return YELLOW_FILL
    return RED_FILL


def _write_metric_block(ws, start_row: int, metrics: list[tuple[str, Any]]) -> int:
    ws.cell(row=start_row, column=1, value="Metric").font = HEADER_FONT
    ws.cell(row=start_row, column=2, value="Value").font = HEADER_FONT
    ws.cell(row=start_row, column=1).fill = HEADER_FILL
    ws.cell(row=start_row, column=2).fill = HEADER_FILL
    row = start_row + 1
    for label, value in metrics:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1
    return row


def _create_dashboard_sheet(
    wb: Workbook,
    manufacturer: str,
    dashboard: dict[str, Any],
    settings: AuditSettings | None = None,
    executive: Any | None = None,
) -> None:
    ws = wb.active
    ws.title = "Dashboard"

    ws["A1"] = "ProTech SI Library Audit Dashboard"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Manufacturer: {manufacturer}"

    metrics = [
        ("Expected MC Deliverables", dashboard["Expected MC Deliverables"]),
        ("Exact Filename Matches", dashboard["Exact Filename Matches"]),
        ("Split Files Present", dashboard["Split Files Present"]),
        ("Passing Compliance", dashboard["Passing Compliance"]),
        ("Needs SI/PCS Review", dashboard["Needs SI/PCS Review"]),
        ("Compliance Rate (%)", dashboard["Compliance Rate (%)"]),
    ]
    if executive:
        metrics.append(("Average Match Confidence", executive.dashboard_extensions.get("Average Match Confidence", 0)))

    row = _write_metric_block(ws, 4, metrics)
    for metric_row in range(5, row):
        label = ws.cell(row=metric_row, column=1).value
        if label == "Compliance Rate (%)":
            ws.cell(row=metric_row, column=2).fill = _compliance_tone(
                float(ws.cell(row=metric_row, column=2).value or 0),
                settings,
            )

    if executive:
        pcs_start = row + 2
        ws.cell(row=pcs_start, column=1, value="PCS Failure Analysis").font = SECTION_FONT
        pcs_metrics = [
            ("Total Expected MC Deliverables", executive.pcs_defense["Total Expected MC Deliverables"]),
            ("Exact Matches", executive.pcs_defense["Exact Matches"]),
            ("Split Files Present", executive.pcs_defense["Split Files Present"]),
            ("Missing From OneDrive Export", executive.pcs_defense["Missing From OneDrive Export"]),
            ("Potential Classification/Naming Mismatch", executive.pcs_defense["Potential Classification/Naming Mismatch"]),
            ("Split File Naming Difference", executive.pcs_defense["Split File Naming Difference"]),
            ("Estimated True Missing Files", executive.pcs_defense["Estimated True Missing Files"]),
            ("Estimated PCS / Mapping Review Items", executive.pcs_defense["Estimated PCS / Mapping Review Items"]),
        ]
        row = _write_metric_block(ws, pcs_start + 1, pcs_metrics)

        rule_start = row + 2
        ws.cell(row=rule_start, column=1, value="Adaptive Rule Impact").font = SECTION_FONT
        row = _write_dataframe(ws, executive.rule_impact, start_row=rule_start + 1)

        summary_start = row + 2
        ws.cell(row=summary_start, column=1, value="Executive Summary").font = SECTION_FONT
        ws.cell(row=summary_start + 1, column=1, value=executive.executive_summary)
        ws.cell(row=summary_start + 1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        row = summary_start + 3

    status_df: pd.DataFrame = dashboard["Status Breakdown"]
    start_status = row + 1
    ws.cell(row=start_status, column=1, value="Status Breakdown").font = SECTION_FONT
    row = _write_dataframe(ws, status_df, start_row=start_status + 1)

    clusters_df: pd.DataFrame = dashboard["Top Missing Clusters"]
    start_clusters = row + 2
    ws.cell(row=start_clusters, column=1, value="Top Missing Clusters").font = SECTION_FONT
    _write_dataframe(ws, clusters_df, start_row=start_clusters + 1)

    _autosize_columns(ws)


def _create_method_notes_sheet(wb: Workbook, notes: list[str]) -> None:
    ws = wb.create_sheet("Method Notes")
    ws["A1"] = "Method Notes"
    ws["A1"].font = TITLE_FONT
    for idx, line in enumerate(notes, start=3):
        ws.cell(row=idx, column=1, value=line)
    ws.column_dimensions["A"].width = 100
    for row in range(3, 3 + len(notes)):
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")


def _create_analysis_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame, title: str) -> None:
    ws = wb.create_sheet(sheet_name)
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    _write_dataframe(ws, df, start_row=3)
    _autosize_columns(ws)


def write_audit_workbook(
    result: AuditResult,
    settings: AuditSettings | None = None,
    executive: Any | None = None,
) -> BytesIO:
    """Build the full audit workbook and return it as an in-memory byte stream."""
    if executive is None and build_executive_package is not None:
        executive = build_executive_package(result)

    wb = Workbook()
    _create_dashboard_sheet(wb, result.manufacturer, result.dashboard, settings, executive)

    audit_df = executive.enriched_audit_df.copy() if executive else result.audit_df.copy()
    missing_df = audit_df[audit_df["Match Status"] == MATCH_MISSING].copy()
    needs_review_df = audit_df[
        audit_df["Match Status"].isin(
            [MATCH_SPLIT_NAMING, MATCH_POTENTIAL_MISMATCH, MATCH_SPLIT_PRESENT]
        )
    ].copy()

    onedrive_inventory = result.onedrive_df[["Name", "Item Type", "Source Sheet"]].copy()
    if "Path" in result.onedrive_df.columns:
        onedrive_inventory["Path"] = result.onedrive_df["Path"]
    expected_from_mc = result.expected_df.copy()

    sheet_specs = [
        ("Audit Results", audit_df),
        ("Missing Files", missing_df),
        ("Needs Review", needs_review_df),
        ("OneDrive Inventory", onedrive_inventory),
        ("Expected From MC", expected_from_mc),
    ]

    for sheet_name, df in sheet_specs:
        ws = wb.create_sheet(sheet_name)
        last_row = _write_dataframe(ws, df)
        if sheet_name in {"Audit Results", "Missing Files", "Needs Review"} and not df.empty:
            status_col = list(df.columns).index("Match Status") + 1
            owner_col = list(df.columns).index("Issue Owner") + 1
            _apply_audit_formatting(ws, status_col, owner_col, last_row - 1)
        _autosize_columns(ws)

    if executive:
        _create_analysis_sheet(
            wb,
            "System Analysis",
            executive.system_analysis,
            "System Failure Analysis",
        )
        _create_analysis_sheet(
            wb,
            "Vehicle Analysis",
            executive.vehicle_analysis,
            "Vehicle Family Analysis",
        )
        root_cause = executive.dashboard_extensions.get("Root Cause Breakdown")
        if root_cause is not None:
            _create_analysis_sheet(wb, "Root Cause Analysis", root_cause, "Predicted Root Cause Breakdown")

    _create_method_notes_sheet(wb, result.method_notes)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_output_filename(manufacturer: str) -> str:
    safe_name = "".join(ch if ch.isalnum() or ch in (" ", "-", "_") else "" for ch in manufacturer)
    safe_name = safe_name.strip().replace(" ", "_") or "Manufacturer"
    return f"{safe_name}_SI_Library_Audit.xlsx"
