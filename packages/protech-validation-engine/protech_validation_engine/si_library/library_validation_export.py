"""Excel export for Library Validation Center reports."""

from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from .library_validation import (
    CLASS_CLASSIFICATION_REVIEW,
    CLASS_LIBRARY_EXISTS,
    CLASS_MANUAL_REVIEW,
    CLASS_NAMING_REVIEW,
    CLASS_NO_RECORD,
    CLASS_PCS_MAPPING,
    CLASS_TRUE_MISSING,
    LibraryValidationResult,
)

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BLUE_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F5F99", end_color="1F5F99", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14, color="0F2F57")
SECTION_FONT = Font(bold=True, size=12, color="0F2F57")
LEGEND_FONT = Font(bold=True, size=11, color="0F2F57")

EXPORT_FILENAME = "Library_Validation_Report.xlsx"

CLASSIFICATION_FILLS: dict[str, PatternFill] = {
    CLASS_LIBRARY_EXISTS: GREEN_FILL,
    CLASS_PCS_MAPPING: YELLOW_FILL,
    CLASS_TRUE_MISSING: RED_FILL,
    CLASS_NAMING_REVIEW: YELLOW_FILL,
    CLASS_CLASSIFICATION_REVIEW: ORANGE_FILL,
    CLASS_NO_RECORD: ORANGE_FILL,
    CLASS_MANUAL_REVIEW: ORANGE_FILL,
}

SUMMARY_ANSWER_FILLS: list[tuple[str, PatternFill]] = [
    ("appear compliant", GREEN_FILL),
    ("true library issues", RED_FILL),
    ("pcs/mapping", YELLOW_FILL),
    ("naming review", YELLOW_FILL),
    ("classification review", ORANGE_FILL),
    ("no matching", ORANGE_FILL),
    ("manual review", ORANGE_FILL),
    ("matched to latest audit", BLUE_FILL),
]


def _autosize_columns(ws, max_width: int = 60) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_length + 2, 12), max_width)


def _write_dataframe(ws, df: pd.DataFrame, start_row: int = 1) -> int:
    for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    if not df.empty:
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if df.empty:
        return start_row
    return start_row + len(df)


def _color_rows_by_classification(ws, df: pd.DataFrame, start_row: int) -> None:
    if df.empty or "Validation Classification" not in df.columns:
        return

    class_col = list(df.columns).index("Validation Classification") + 1
    col_count = len(df.columns)
    first_data_row = start_row + 1
    last_data_row = start_row + len(df)
    color_full_row = len(df) <= 2500

    for row_idx in range(first_data_row, last_data_row + 1):
        class_value = str(ws.cell(row=row_idx, column=class_col).value or "")
        row_fill = CLASSIFICATION_FILLS.get(class_value)
        if not row_fill:
            continue
        target_cols = range(1, col_count + 1) if color_full_row else (class_col,)
        for col_idx in target_cols:
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    class_letter = get_column_letter(class_col)
    class_range = f"{class_letter}{first_data_row}:{class_letter}{last_data_row}"
    for class_name, fill in CLASSIFICATION_FILLS.items():
        ws.conditional_formatting.add(
            class_range,
            CellIsRule(operator="equal", formula=[f'"{class_name}"'], fill=fill),
        )


def _color_executive_answers(ws, start_row: int, row_count: int) -> None:
    for row_idx in range(start_row + 1, start_row + row_count + 1):
        question = str(ws.cell(row=row_idx, column=1).value or "").lower()
        answer_cell = ws.cell(row=row_idx, column=2)
        for keyword, fill in SUMMARY_ANSWER_FILLS:
            if keyword in question:
                answer_cell.fill = fill
                break
        answer_cell.alignment = Alignment(vertical="top", wrap_text=True)


def _write_validation_sheet(ws, df: pd.DataFrame, start_row: int = 1) -> None:
    last_row = _write_dataframe(ws, df, start_row=start_row)
    _color_rows_by_classification(ws, df, start_row)
    if not df.empty and "Confidence" in df.columns:
        conf_col = list(df.columns).index("Confidence") + 1
        for row_idx in range(start_row + 1, last_row + 1):
            cell = ws.cell(row=row_idx, column=conf_col)
            try:
                score = float(cell.value)
            except (TypeError, ValueError):
                continue
            if score >= 95:
                cell.fill = GREEN_FILL
            elif score >= 60:
                cell.fill = YELLOW_FILL
            elif score > 0:
                cell.fill = ORANGE_FILL
            else:
                cell.fill = RED_FILL


def _write_legend(ws, start_row: int) -> None:
    ws.cell(row=start_row, column=1, value="Color Legend").font = LEGEND_FONT
    legend_rows = [
        (CLASS_LIBRARY_EXISTS, GREEN_FILL, "Good — library file appears present"),
        (CLASS_PCS_MAPPING, YELLOW_FILL, "Review — likely PCS attachment or mapping"),
        (CLASS_TRUE_MISSING, RED_FILL, "Bad — true missing from audited library"),
        (CLASS_NAMING_REVIEW, YELLOW_FILL, "Review — naming difference"),
        (CLASS_CLASSIFICATION_REVIEW, ORANGE_FILL, "Review — classification mismatch"),
        (CLASS_NO_RECORD, ORANGE_FILL, "Review — no matching MC expected record"),
        (CLASS_MANUAL_REVIEW, ORANGE_FILL, "Review — manual follow-up required"),
    ]
    row = start_row + 1
    for label, fill, note in legend_rows:
        ws.cell(row=row, column=1, value=label).fill = fill
        ws.cell(row=row, column=2, value=note)
        row += 1


def _executive_summary_sections(result: LibraryValidationResult) -> list[tuple[str, pd.DataFrame]]:
    summary = result.summary
    overview = pd.DataFrame(
        [
            {"Question": "How many records were validated?", "Answer": summary.records_validated},
            {"Question": "How many appear compliant (library file exists)?", "Answer": summary.appear_compliant},
            {
                "Question": "How many appear to be true library issues?",
                "Answer": summary.true_missing,
            },
            {
                "Question": "How many appear to be PCS/mapping issues?",
                "Answer": summary.pcs_mapping_review,
            },
            {"Question": "Rows matched to latest audit database", "Answer": summary.rows_matched_to_audit},
            {"Question": "Naming review required", "Answer": summary.naming_review},
            {"Question": "Classification review required", "Answer": summary.classification_review},
            {"Question": "No matching expected record", "Answer": summary.no_matching_record},
            {"Question": "Manual review required", "Answer": summary.manual_review},
            {"Question": "Source file", "Answer": result.source_filename},
        ]
    )
    sections = [("Executive Overview", overview)]
    if not summary.top_systems.empty:
        sections.append(("Most Affected Systems", summary.top_systems))
    if not summary.top_models.empty:
        sections.append(("Most Affected Models", summary.top_models))
    return sections


def write_library_validation_workbook(result: LibraryValidationResult) -> BytesIO:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    ws_summary["A1"] = "Library Validation Center Report"
    ws_summary["A1"].font = TITLE_FONT
    ws_summary["A2"] = (
        "Validate external exception reports against the latest audited SI Library. "
        "This analysis does not replace the audit engine. It estimates whether each "
        "uploaded record appears to be a library problem, PCS/mapping issue, or manual review item."
    )
    ws_summary["A2"].alignment = Alignment(wrap_text=True)

    row = 4
    for title, frame in _executive_summary_sections(result):
        ws_summary.cell(row=row, column=1, value=title).font = SECTION_FONT
        table_start = row + 1
        row = _write_dataframe(ws_summary, frame, start_row=table_start) + 1
        if title == "Executive Overview" and not frame.empty:
            _color_executive_answers(ws_summary, table_start, len(frame))

    legend_row = row + 1
    _write_legend(ws_summary, legend_row)
    _autosize_columns(ws_summary)

    detailed = result.detailed_results
    display_cols = [col for col in detailed.columns if col not in {"Source Row", "Original Row Data"}]

    manual_mask = detailed["Validation Classification"].isin(
        {CLASS_MANUAL_REVIEW, CLASS_NO_RECORD, CLASS_CLASSIFICATION_REVIEW}
    )

    sheets: list[tuple[str, pd.DataFrame]] = [
        ("Validation Results", detailed[display_cols] if display_cols else detailed),
        ("True Missing Files", detailed[detailed["Validation Classification"] == CLASS_TRUE_MISSING][display_cols]),
        ("PCS Mapping Review", detailed[detailed["Validation Classification"] == CLASS_PCS_MAPPING][display_cols]),
        ("Naming Review", detailed[detailed["Validation Classification"] == CLASS_NAMING_REVIEW][display_cols]),
        ("Manual Review", detailed[manual_mask][display_cols]),
        ("Source Upload", result.source_upload),
    ]

    for title, frame in sheets:
        ws = wb.create_sheet(title)
        sheet_df = frame if not frame.empty else pd.DataFrame(columns=display_cols)
        if title == "Source Upload":
            _write_dataframe(ws, sheet_df)
        else:
            _write_validation_sheet(ws, sheet_df)
        _autosize_columns(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
