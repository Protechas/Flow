"""QA Intelligence Engine — Phase 1 (rules-based, no AI).

Takes a set of uploaded workbooks (an MC chart plus reference sheets), parses
and normalizes them, runs structured checks, and produces findings plus a
formatted multi-sheet Excel report.

Checks:
  missing_required_column  — a chart-like sheet lacks year/make/model/system
  blank_cell               — blanks in important (mostly-filled) columns
  duplicate_row            — exact duplicate rows
  duplicate_component      — same year/make/model/system appearing twice
  inconsistent_manufacturer— casing/spacing variants of the same make
  malformed_value          — years that aren't plausible 4-digit numbers,
                             whitespace-padded cells, literal "nan"
  conflicting_data         — same component key, different values in a file
  cross_file_mismatch      — reference sheet entries missing from the MC
                             chart, or value conflicts against it

Designed to stay modular: each check is a function returning findings, so
future phases can add document parsing, AI summaries, and task creation.
"""

from __future__ import annotations

import io
import re
from collections import defaultdict
from dataclasses import dataclass, field

import pandas as pd

REQUIRED_CHART_COLUMNS = ["year", "make", "model", "system"]
KEY_COLUMNS = ["year", "make", "model", "system"]

SEVERITY_ORDER = {"high": 0, "medium": 1, "low": 2}

PRIORITY_BY_SEVERITY = {"high": "High", "medium": "Medium", "low": "Low"}


@dataclass
class QaFinding:
    issue_type: str
    severity: str
    title: str
    source_file: str
    sheet_name: str
    row_number: int | None
    column_name: str | None
    expected: str
    found: str
    explanation: str
    suggested_task_title: str
    suggested_task_description: str
    suggested_priority: str
    suggested_assignee: str = "SI Team"

    def to_dict(self) -> dict:
        return {
            "issue_type": self.issue_type,
            "severity": self.severity,
            "title": self.title,
            "source_file": self.source_file,
            "sheet_name": self.sheet_name,
            "row_number": self.row_number,
            "column_name": self.column_name,
            "expected": self.expected,
            "found": self.found,
            "explanation": self.explanation,
            "suggested_task_title": self.suggested_task_title,
            "suggested_task_description": self.suggested_task_description,
            "suggested_priority": self.suggested_priority,
            "suggested_assignee": self.suggested_assignee,
        }


@dataclass
class ParsedSheet:
    source_file: str
    sheet_name: str
    frame: pd.DataFrame
    is_chart: bool = False


@dataclass
class QaScanResult:
    files_scanned: int
    sheets_scanned: int
    rows_scanned: int
    findings: list[QaFinding] = field(default_factory=list)


def _norm(value) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = str(value).strip()
    if text.lower() == "nan":
        return ""
    return re.sub(r"\s+", " ", text)


def _canon(value) -> str:
    return _norm(value).lower()


def _parse_files(files: list[dict]) -> list[ParsedSheet]:
    sheets: list[ParsedSheet] = []
    for f in files:
        name = str(f.get("name") or "workbook.xlsx")
        data: bytes = f["bytes"]
        is_chart = bool(f.get("is_chart"))
        try:
            if name.lower().endswith(".csv"):
                frame = pd.read_csv(io.BytesIO(data))
                frame.columns = [str(c).strip() for c in frame.columns]
                sheets.append(ParsedSheet(name, "Sheet1", frame, is_chart))
                continue
            xl = pd.ExcelFile(io.BytesIO(data))
            for sheet_name in xl.sheet_names:
                frame = xl.parse(sheet_name)
                if frame.empty:
                    continue
                frame.columns = [str(c).strip() for c in frame.columns]
                sheets.append(ParsedSheet(name, sheet_name, frame, is_chart))
        except Exception as exc:  # noqa: BLE001 — a bad file is a finding, not a crash
            sheets.append(
                ParsedSheet(name, "(unreadable)", pd.DataFrame({"__error__": [str(exc)]}), is_chart)
            )
    return sheets


def _col_lookup(frame: pd.DataFrame) -> dict[str, str]:
    return {str(c).strip().lower(): c for c in frame.columns}


def _key_for_row(row, lookup: dict[str, str]) -> tuple | None:
    parts = tuple(_canon(row.get(lookup[k])) for k in KEY_COLUMNS if k in lookup)
    if len(parts) < 2 or not any(parts):
        return None
    return parts


def _display_key(row, lookup: dict[str, str]) -> str:
    return " ".join(
        _norm(row.get(lookup[k])) for k in KEY_COLUMNS if k in lookup and _norm(row.get(lookup[k]))
    )


# ——— Checks ————————————————————————————————————————————————————————————————


def check_unreadable(sheet: ParsedSheet) -> list[QaFinding]:
    if "__error__" not in sheet.frame.columns:
        return []
    err = str(sheet.frame["__error__"].iloc[0])
    return [
        QaFinding(
            issue_type="unreadable_file",
            severity="high",
            title=f"Could not read {sheet.source_file}",
            source_file=sheet.source_file,
            sheet_name=sheet.sheet_name,
            row_number=None,
            column_name=None,
            expected="A readable Excel/CSV file",
            found=err[:200],
            explanation="The file could not be parsed — it may be corrupted or in an unsupported format.",
            suggested_task_title=f"Re-export {sheet.source_file}",
            suggested_task_description="Open the source file, verify it opens cleanly, and re-export it as .xlsx.",
            suggested_priority="High",
        )
    ]


def check_required_columns(sheet: ParsedSheet) -> list[QaFinding]:
    if not sheet.is_chart:
        return []
    lookup = _col_lookup(sheet.frame)
    findings = []
    for col in REQUIRED_CHART_COLUMNS:
        if col not in lookup:
            findings.append(
                QaFinding(
                    issue_type="missing_required_column",
                    severity="high",
                    title=f"MC chart is missing the '{col.title()}' column",
                    source_file=sheet.source_file,
                    sheet_name=sheet.sheet_name,
                    row_number=None,
                    column_name=col.title(),
                    expected=f"A '{col.title()}' column",
                    found=f"Columns present: {', '.join(map(str, sheet.frame.columns[:10]))}",
                    explanation="Manufacturer charts need Year/Make/Model/System columns for downstream validation.",
                    suggested_task_title=f"Add '{col.title()}' column to {sheet.source_file}",
                    suggested_task_description=f"Update the chart '{sheet.sheet_name}' so every record has a {col.title()} value.",
                    suggested_priority="High",
                )
            )
    return findings


def check_blank_cells(sheet: ParsedSheet) -> list[QaFinding]:
    frame = sheet.frame
    if frame.empty or "__error__" in frame.columns:
        return []
    findings = []
    lookup = _col_lookup(frame)
    important = {lookup[k] for k in REQUIRED_CHART_COLUMNS if k in lookup}
    # Any column that's ≥60% filled counts as important too.
    for col in frame.columns:
        filled = frame[col].map(lambda v: bool(_norm(v))).mean()
        if filled >= 0.6:
            important.add(col)
    for col in important:
        blanks = [
            int(idx) + 2
            for idx, v in frame[col].items()
            if not _norm(v)
        ]
        if not blanks or len(blanks) == len(frame):
            continue
        rows_label = ", ".join(map(str, blanks[:12])) + ("…" if len(blanks) > 12 else "")
        findings.append(
            QaFinding(
                issue_type="blank_cell",
                severity="medium" if col in {lookup.get(k) for k in REQUIRED_CHART_COLUMNS} else "low",
                title=f"{len(blanks)} blank cell(s) in '{col}'",
                source_file=sheet.source_file,
                sheet_name=sheet.sheet_name,
                row_number=blanks[0],
                column_name=str(col),
                expected="A value in every row of this column",
                found=f"Blank at rows {rows_label}",
                explanation="This column is filled for most rows, so blanks are likely omissions.",
                suggested_task_title=f"Fill blank '{col}' values in {sheet.source_file}",
                suggested_task_description=f"Sheet '{sheet.sheet_name}': fill the missing '{col}' values at rows {rows_label}.",
                suggested_priority="Medium",
            )
        )
    return findings


def check_duplicates(sheet: ParsedSheet) -> list[QaFinding]:
    frame = sheet.frame
    if frame.empty or "__error__" in frame.columns:
        return []
    findings = []
    normalized = frame.map(_canon)
    dup_mask = normalized.duplicated(keep="first")
    dup_rows = [int(i) + 2 for i in frame.index[dup_mask]]
    if dup_rows:
        rows_label = ", ".join(map(str, dup_rows[:12])) + ("…" if len(dup_rows) > 12 else "")
        findings.append(
            QaFinding(
                issue_type="duplicate_row",
                severity="medium",
                title=f"{len(dup_rows)} exact duplicate row(s)",
                source_file=sheet.source_file,
                sheet_name=sheet.sheet_name,
                row_number=dup_rows[0],
                column_name=None,
                expected="Each row to appear once",
                found=f"Duplicates at rows {rows_label}",
                explanation="Identical rows inflate counts and usually mean copy-paste artifacts.",
                suggested_task_title=f"Remove duplicate rows from {sheet.source_file}",
                suggested_task_description=f"Sheet '{sheet.sheet_name}': delete the duplicate rows at {rows_label}.",
                suggested_priority="Medium",
            )
        )

    lookup = _col_lookup(frame)
    if all(k in lookup for k in ("make", "model", "system")):
        seen: dict[tuple, int] = {}
        for idx, row in frame.iterrows():
            key = _key_for_row(row, lookup)
            if key is None:
                continue
            if key in seen and not dup_mask.get(idx, False):
                findings.append(
                    QaFinding(
                        issue_type="duplicate_component",
                        severity="medium",
                        title=f"Duplicate component: {_display_key(row, lookup)}",
                        source_file=sheet.source_file,
                        sheet_name=sheet.sheet_name,
                        row_number=int(idx) + 2,
                        column_name=None,
                        expected="One row per component",
                        found=f"Also defined at row {seen[key]}",
                        explanation="The same Year/Make/Model/System appears more than once with differing data elsewhere in the row.",
                        suggested_task_title=f"Consolidate duplicate {_display_key(row, lookup)}",
                        suggested_task_description=f"Sheet '{sheet.sheet_name}': merge rows {seen[key]} and {int(idx) + 2} into one record.",
                        suggested_priority="Medium",
                    )
                )
            else:
                seen.setdefault(key, int(idx) + 2)
    return findings


def check_manufacturer_consistency(sheets: list[ParsedSheet]) -> list[QaFinding]:
    variants: dict[str, dict[str, tuple[str, str, int]]] = defaultdict(dict)
    for sheet in sheets:
        lookup = _col_lookup(sheet.frame)
        make_col = lookup.get("make") or lookup.get("manufacturer")
        if not make_col:
            continue
        for idx, value in sheet.frame[make_col].items():
            raw = _norm(value)
            if not raw:
                continue
            canon = re.sub(r"[^a-z0-9]", "", raw.lower())
            variants[canon].setdefault(raw, (sheet.source_file, sheet.sheet_name, int(idx) + 2))
    findings = []
    for canon, spellings in variants.items():
        if len(spellings) < 2:
            continue
        names = sorted(spellings.keys())
        first_file, first_sheet, first_row = spellings[names[1]]
        findings.append(
            QaFinding(
                issue_type="inconsistent_manufacturer",
                severity="medium",
                title=f"Manufacturer spelled {len(names)} ways: {', '.join(names[:4])}",
                source_file=first_file,
                sheet_name=first_sheet,
                row_number=first_row,
                column_name="Make",
                expected=f"One consistent spelling (e.g. '{names[0]}')",
                found=" / ".join(names[:6]),
                explanation="Different spellings of the same manufacturer break matching between sheets.",
                suggested_task_title=f"Standardize manufacturer name '{names[0]}'",
                suggested_task_description=f"Pick one spelling and update all files: {', '.join(names)}.",
                suggested_priority="Medium",
            )
        )
    return findings


def check_malformed_values(sheet: ParsedSheet) -> list[QaFinding]:
    frame = sheet.frame
    if frame.empty or "__error__" in frame.columns:
        return []
    findings = []
    lookup = _col_lookup(frame)
    year_col = lookup.get("year")
    if year_col:
        bad = []
        for idx, v in frame[year_col].items():
            text = _norm(v)
            if not text:
                continue
            digits = re.sub(r"\.0$", "", text)
            if not re.fullmatch(r"(19|20)\d{2}", digits):
                bad.append((int(idx) + 2, text))
        for row_num, text in bad[:20]:
            findings.append(
                QaFinding(
                    issue_type="malformed_value",
                    severity="medium",
                    title=f"Invalid year '{text}'",
                    source_file=sheet.source_file,
                    sheet_name=sheet.sheet_name,
                    row_number=row_num,
                    column_name=str(year_col),
                    expected="A 4-digit year (1900–2099)",
                    found=text,
                    explanation="Year values that aren't plausible 4-digit years break sorting and matching.",
                    suggested_task_title=f"Fix year at row {row_num} in {sheet.source_file}",
                    suggested_task_description=f"Sheet '{sheet.sheet_name}': correct '{text}' to a valid 4-digit year.",
                    suggested_priority="Medium",
                )
            )
    for col in frame.columns:
        padded = [
            (int(idx) + 2, str(v))
            for idx, v in frame[col].items()
            if isinstance(v, str) and v != v.strip() and v.strip()
        ]
        if padded:
            rows_label = ", ".join(str(r) for r, _ in padded[:10]) + ("…" if len(padded) > 10 else "")
            findings.append(
                QaFinding(
                    issue_type="malformed_value",
                    severity="low",
                    title=f"{len(padded)} value(s) with stray whitespace in '{col}'",
                    source_file=sheet.source_file,
                    sheet_name=sheet.sheet_name,
                    row_number=padded[0][0],
                    column_name=str(col),
                    expected="Values without leading/trailing spaces",
                    found=f"Rows {rows_label}",
                    explanation="Padded values look identical but fail exact matching.",
                    suggested_task_title=f"Trim whitespace in '{col}' ({sheet.source_file})",
                    suggested_task_description=f"Sheet '{sheet.sheet_name}': remove leading/trailing spaces at rows {rows_label}.",
                    suggested_priority="Low",
                )
            )
    return findings


def check_conflicting_data(sheet: ParsedSheet) -> list[QaFinding]:
    frame = sheet.frame
    if frame.empty or "__error__" in frame.columns:
        return []
    lookup = _col_lookup(frame)
    if not all(k in lookup for k in ("make", "model", "system")):
        return []
    key_cols = [lookup[k] for k in KEY_COLUMNS if k in lookup]
    value_cols = [c for c in frame.columns if c not in key_cols]
    if not value_cols:
        return []
    findings = []
    by_key: dict[tuple, tuple[int, pd.Series]] = {}
    for idx, row in frame.iterrows():
        key = _key_for_row(row, lookup)
        if key is None:
            continue
        if key in by_key:
            first_row_num, first = by_key[key]
            diffs = [
                (str(c), _norm(first.get(c)), _norm(row.get(c)))
                for c in value_cols
                if _canon(first.get(c)) != _canon(row.get(c))
                and (_norm(first.get(c)) or _norm(row.get(c)))
            ]
            if diffs:
                col, a, b = diffs[0]
                findings.append(
                    QaFinding(
                        issue_type="conflicting_data",
                        severity="high",
                        title=f"Conflicting data for {_display_key(row, lookup)}",
                        source_file=sheet.source_file,
                        sheet_name=sheet.sheet_name,
                        row_number=int(idx) + 2,
                        column_name=col,
                        expected=f"Consistent values (row {first_row_num}: '{a}')",
                        found=f"Row {int(idx) + 2}: '{b}'" + (f" (+{len(diffs) - 1} more columns)" if len(diffs) > 1 else ""),
                        explanation="The same component is defined twice with different data — one of them is wrong.",
                        suggested_task_title=f"Resolve conflicting data for {_display_key(row, lookup)}",
                        suggested_task_description=f"Sheet '{sheet.sheet_name}': rows {first_row_num} and {int(idx) + 2} disagree on {', '.join(d[0] for d in diffs)}. Confirm the correct values and remove the other row.",
                        suggested_priority="High",
                    )
                )
        else:
            by_key[key] = (int(idx) + 2, row)
    return findings


def check_cross_file(chart_sheets: list[ParsedSheet], other_sheets: list[ParsedSheet]) -> list[QaFinding]:
    chart_keys: dict[tuple, tuple[str, pd.Series, dict[str, str]]] = {}
    for sheet in chart_sheets:
        lookup = _col_lookup(sheet.frame)
        if not all(k in lookup for k in ("make", "model", "system")):
            continue
        for _, row in sheet.frame.iterrows():
            key = _key_for_row(row, lookup)
            if key is not None:
                chart_keys.setdefault(key, (sheet.source_file, row, lookup))
    if not chart_keys:
        return []
    findings = []
    for sheet in other_sheets:
        lookup = _col_lookup(sheet.frame)
        if not all(k in lookup for k in ("make", "model", "system")):
            continue
        for idx, row in sheet.frame.iterrows():
            key = _key_for_row(row, lookup)
            if key is None:
                continue
            if key not in chart_keys:
                findings.append(
                    QaFinding(
                        issue_type="cross_file_mismatch",
                        severity="high",
                        title=f"Not in MC chart: {_display_key(row, lookup)}",
                        source_file=sheet.source_file,
                        sheet_name=sheet.sheet_name,
                        row_number=int(idx) + 2,
                        column_name=None,
                        expected="A matching record in the MC chart",
                        found="No chart record with this Year/Make/Model/System",
                        explanation="This reference sheet lists a component the MC chart doesn't have.",
                        suggested_task_title=f"Reconcile {_display_key(row, lookup)} with the MC chart",
                        suggested_task_description=f"'{sheet.source_file}' row {int(idx) + 2} has no matching MC chart record — add it to the chart or remove it from the reference sheet.",
                        suggested_priority="High",
                    )
                )
                continue
            chart_file, chart_row, chart_lookup = chart_keys[key]
            shared = [
                c
                for c in _col_lookup(sheet.frame)
                if c in chart_lookup and c not in KEY_COLUMNS
            ]
            for c in shared:
                ours = _canon(row.get(lookup[c]))
                theirs = _canon(chart_row.get(chart_lookup[c]))
                if ours and theirs and ours != theirs:
                    findings.append(
                        QaFinding(
                            issue_type="cross_file_mismatch",
                            severity="medium",
                            title=f"'{lookup[c]}' disagrees with the MC chart: {_display_key(row, lookup)}",
                            source_file=sheet.source_file,
                            sheet_name=sheet.sheet_name,
                            row_number=int(idx) + 2,
                            column_name=str(lookup[c]),
                            expected=f"Chart value '{_norm(chart_row.get(chart_lookup[c]))}' ({chart_file})",
                            found=_norm(row.get(lookup[c])),
                            explanation="The reference sheet and the MC chart disagree — one of them is stale.",
                            suggested_task_title=f"Align '{lookup[c]}' for {_display_key(row, lookup)}",
                            suggested_task_description=f"Compare '{sheet.source_file}' row {int(idx) + 2} against the MC chart and update the wrong side.",
                            suggested_priority="Medium",
                        )
                    )
    return findings


# ——— Orchestration ——————————————————————————————————————————————————————————


def run_qa_scan(files: list[dict]) -> QaScanResult:
    sheets = _parse_files(files)
    chart_sheets = [s for s in sheets if s.is_chart and "__error__" not in s.frame.columns]
    other_sheets = [s for s in sheets if not s.is_chart and "__error__" not in s.frame.columns]

    findings: list[QaFinding] = []
    for sheet in sheets:
        findings += check_unreadable(sheet)
        findings += check_required_columns(sheet)
        findings += check_blank_cells(sheet)
        findings += check_duplicates(sheet)
        findings += check_malformed_values(sheet)
        findings += check_conflicting_data(sheet)
    findings += check_manufacturer_consistency(sheets)
    findings += check_cross_file(chart_sheets, other_sheets)

    findings.sort(key=lambda f: (SEVERITY_ORDER.get(f.severity, 3), f.source_file, f.issue_type))
    rows = sum(len(s.frame) for s in sheets if "__error__" not in s.frame.columns)
    return QaScanResult(
        files_scanned=len({s.source_file for s in sheets}),
        sheets_scanned=len(sheets),
        rows_scanned=rows,
        findings=findings,
    )


# ——— Excel report ————————————————————————————————————————————————————————————

REPORT_COLUMNS = [
    ("Severity", "severity"),
    ("Issue Type", "issue_type"),
    ("Title", "title"),
    ("Source File", "source_file"),
    ("Sheet", "sheet_name"),
    ("Row", "row_number"),
    ("Column", "column_name"),
    ("Expected", "expected"),
    ("Found", "found"),
    ("Explanation", "explanation"),
    ("Suggested Task", "suggested_task_title"),
    ("Task Details", "suggested_task_description"),
    ("Priority", "suggested_priority"),
    ("Assignee/Team", "suggested_assignee"),
    ("Status", "status"),
]


def _findings_frame(findings: list[dict]) -> pd.DataFrame:
    rows = []
    for f in findings:
        rows.append({label: f.get(key, "") for label, key in REPORT_COLUMNS})
    return pd.DataFrame(rows, columns=[label for label, _ in REPORT_COLUMNS])


def _style_sheet(writer, sheet_name: str, frame: pd.DataFrame) -> None:
    from openpyxl.styles import Font, PatternFill

    ws = writer.sheets[sheet_name]
    ws.freeze_panes = "A2"
    if frame.shape[0] > 0:
        ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="1F2A44")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for col_idx, column in enumerate(frame.columns, start=1):
        sample = [str(column)] + [str(v) for v in frame[column].head(40)]
        width = min(max(len(s) for s in sample) + 2, 60)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width


def write_qa_report(result_summary: dict, findings: list[dict]) -> io.BytesIO:
    buffer = io.BytesIO()
    frame = _findings_frame(findings)
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Summary first — the leadership view.
        by = lambda key: (  # noqa: E731
            pd.Series([f.get(key) or "—" for f in findings]).value_counts().rename_axis(key.replace("_", " ").title()).reset_index(name="Findings")
        )
        summary_rows = [
            {"Metric": "Files scanned", "Value": result_summary.get("files_scanned", 0)},
            {"Metric": "Sheets scanned", "Value": result_summary.get("sheets_scanned", 0)},
            {"Metric": "Rows scanned", "Value": result_summary.get("rows_scanned", 0)},
            {"Metric": "Total findings", "Value": len(findings)},
        ]
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Summary", index=False, startrow=0)
        start = len(summary_rows) + 3
        for key in ("severity", "issue_type", "source_file", "status"):
            table = by(key)
            table.to_excel(writer, sheet_name="Summary", index=False, startrow=start)
            start += len(table) + 3
        _style_sheet(writer, "Summary", pd.DataFrame(summary_rows))

        frame.to_excel(writer, sheet_name="All Findings", index=False)
        _style_sheet(writer, "All Findings", frame)

        high = frame[frame["Severity"] == "high"]
        high.to_excel(writer, sheet_name="High Severity", index=False)
        _style_sheet(writer, "High Severity", high)

        for sheet_name, sort_col in (
            ("By Source File", "Source File"),
            ("By Issue Type", "Issue Type"),
            ("By Assignee", "Assignee/Team"),
        ):
            grouped = frame.sort_values([sort_col, "Severity"]) if len(frame) else frame
            grouped.to_excel(writer, sheet_name=sheet_name, index=False)
            _style_sheet(writer, sheet_name, grouped)
    buffer.seek(0)
    return buffer
